"""Deze module bevat de logica om sas-modulescripts van Teradata op te halen, deze samen met de datadictionary te
verwerken, en de keymodules (met bijbehorende modulegroepen, moduleversies en features) in de juiste
directory-structuur in het Carkit-project te plaatsen.

.. todo:: aanpassen \n\n \
["dg_i_o_40ana_int.keytable_finr_validatie ", "&inputlocatie. "], \
["dg_i_o_40ana_int.keytable_finr_beljaar_vld ", "&inputlocatie. "], \n\n \
aan de hand van het initialisationscript. Hier staat in waarnaar de keytable verwijst. \
of afspraak maken over welke keytable er gebruikt wordt. OF keytable in modulescripts laten staan als macrovar. \
Hier een oplossing voor bedenken.

.. todo:: voor controleren lineage: \n\n \
    mogelijke verbeteringen: \n\n \
    controleren of INNER QUERIES kolommen gebruikt worden. \n \
    AL DOORGEVOERD: voor alle tokens controleren of het woord voorkomt in keywords. \n\n \
    Anders is het waarschijnlijk dat het een kolom betreft en moet warning gegenereerd worden. \n\n \
    voor alle gevonden kolommen controleren of de kolommen ook daadwerkelijk voorkomen in de brontabellen. \n\n \
    GAAT AUTOMATISCH GOED? :woorden tussen dubbele quotes zien als kolommen.  bijv "type", \
    wat anders gezien wordt als keyword. \n\n \
    check op parameters, het kan voorkomen dat een kolomnaam afhankelijk is van een parameter. Dit kan beter opgelost \
    worden door gebruik te maken van gedefinieerde features zonder parameters op basis van custom features.

Het stappenplan om de directory-structuur van keymodules goed neer te zetten, is:

#. Stap 1: ophalen modulescripts.

    #. ga naar sas-eg en log in op de gridserver. (psasogmdl01.prod.belastingdienst.nl)
    #. run utilities/zipscript.sas (zorg dat je de juiste outputnaam opgeeft (de huidige datum meegeven))
    #. run utilities/export2td.sas (upload de juiste folder naar de juiste tabel (huidige datum 2x meegeven))
    #. ga naar de DST. (psaslstbl01.prod.belastingdienst.nl)
    #. run ophalen_modulescripts(teradatanaam)

#. Stap 2: uploaden van datadictionary.

    #. ga naar folder Q:\\\P9007VFSSW12F\\\LKB_BV\\\BI-A\\\KIT\\\CAR_KIT\\\02_Documentatie
    #. ga naar filezilla en log in op de DST
    #. kopieer het huidige datadictionary naar /srv/data/SIM/input/datadictionaries/ input-folder.

#. Stap 3: Draai load_modules met parameters datadictionary en modulescriptlocatie

    #. script leest datadictionary in en maakt een featureset aan
    #. script leest alle bestanden uit de modulescripts in
    #. script bekijkt voor iedere modulescript of deze is gekoppeld in het datadictionary aan enkele features
    #. script maakt voor iedere feature een outputfile. voor iedere keygroep, modeleversie en modulegroep een folder.
"""

import re
from common import ireplace
from settings import *
from tdquery import td_unzip
from datetime import datetime
from openpyxl import load_workbook
from sharedclasses import SimModule
from sharedclasses import Feature


def ophalen_modulescripts(td_locatie_modulescripts, overwrite=True):
    """Haalt het byte-bestand met sas-modulescripts op van Teradata. Output wordt als zip-bestand
    neergezet in het Carkit-project en vervolgens uitgepakt. De sas-modulescripts bevatten de sas-code om de
    Teradata-tabellen te maken van de verschillende modulegroepen met de bijbehorende features.

    :param td_locatie_modulescripts: een combinatie van datagebied en naam van de Teradata-tabel. Deze tabel is \
    een kopie in bytes van een zip-bestand met sas-modulescripts. Bv. DG_I_O_40ANA_INT.modulescripts_20180306
    :param overwrite: overschrijven van het bestand in de Carkit-directory
    """
    naam = td_locatie_modulescripts.split(".")[1].lower()
    td_unzip(td_locatie_modulescripts, '/srv/data/SIM/{}/input/modulescripts/{}'.format(SIM_versie, naam),
             overwrite=overwrite)


def verwerk_input(modulescripts, datadictionary, locatie_backup=locatie_backup_keymodules,
                  locatie_keywords=locatie_td_keywords):
    """Bouwt vanuit datadictionary (is leidend) en sas-modulescripts nieuwe folders op met keymodules, bv. k1.
    Alle bestaande keymodule-folders worden gekopieerd naar een archieffolder met de naam keymodules gevolgd door
    de backup timestamp.
    Voor de datadictionary gelden de volgende eisen:

    * bovenste rij bevat kolomnamen
    * maximaal 100 kolommen
    * sheet met omschrijvingen etc heet datadictionary
    * sheet bevat geen lege rijen

    De Excel-datadictionary bevat data validation kolommen, waarbij alleen bepaalde waarden zijn toegestaan. Deze
    data validation wordt niet meegenomen bij het uitlezen van de Excel. Openpyxl doet hier niets mee, maar geeft wel
    een waarschuwing. Deze kan worden genegeerd.

    De locatie-parameters worden gezet in :func:`sim.settings`

    :param modulescripts: het zip-bestand met de sas-modulescripts
    :param datadictionary: de Excel datadictionary
    :param locatie_backup: de locatie waar de backup wordt weggeschreven
    :param locatie_keymodules: de locatie waar de nieuwe keymodules worden weggeschreven
    :param locatie_keywords: de locatie van Teradata reserved keywords
    :param locatie_modulescripts: de directory met daarin het zip-bestand met sas-modulescripts
    :param locatie_datadictionaries: de directory met daarin de datadictionary in Excel

    """
    root_modulescripts = os.path.join(locatie_modulescripts, modulescripts)
    print('root',root_modulescripts)
    print('locatie',locatie_modulescripts)
    locatie_datadictionary = os.path.join(locatie_datadictionaries, datadictionary)
    locatie_initscript = os.path.join(root_modulescripts, "initialisation_sim.sas")
    timestamp = str(datetime.now()).replace(" ", "_").replace("-", "").replace(":", "").split(".")[0]

    # moduleversie: sim_k1_modulenaam_v1
    # modulenaam  : deel na kx_

    if not os.path.isfile(locatie_initscript):
        print("kan script initialisation_sim.sas niet vinden")
        return

    if not os.path.isfile(locatie_datadictionary):
        print("kan {} niet vinden".format(locatie_datadictionary))

    # stap 1: backup maken van de keymodules

    # comment: mogelijk beter os.rename ipv os.system ?
    os.system('mv {} {}keymodules_{}'.format(locatie_keymodules, locatie_backup, timestamp))
    os.mkdir(locatie_keymodules)
    print('backup gemaakt op {}keymodules_{}'.format(locatie_backup, timestamp))

    # stap 2: aanmaken featureclass en moduleclass.
    featureset = {}
    modulelijst = []
    with open(locatie_keywords) as inputfile:
        print('keywords ingelezen')
        td_keywordlist = [items.strip().lower() for items in inputfile.readline().split(',')]

    wb = load_workbook(filename=locatie_datadictionary)
    sheet = wb['datadictionary']

    # noodzakelijk dat er in het datadictionary geen lege rijen voorkomen.
    kolomnamen = {}
    for i in range(1, 100):
        if sheet.cell(column=i, row=1).value:
            kolomnamen[sheet.cell(column=i, row=1).value] = i

    for benodigde_kolomnaam in benodigde_kolomnamen_datadictionary:
        if benodigde_kolomnaam not in kolomnamen and benodigde_kolomnaam not in ['finr', 'peildatum']:
            print("kolom {} ontbreekt in datadictionary".format(benodigde_kolomnaam))

    i = 2  # skip header
    featurenaam = ""

    while featurenaam or i == 2:
        feature = Feature()
        feature.verwerk_datadictionary_record(kolomnamen, sheet, i)
        if feature.naam:
            featurenaam = feature.moduleversie + ";" + feature.naam
            featureset[featurenaam] = feature
            if feature.moduleversie not in modulelijst:
                modulelijst.append(feature.moduleversie)
                # print(feature.moduleversie,"toegevoegd aan modulelijst")
        else:
            featurenaam = None
        i += 1
    print("datadictionary verwerkt")

    #
    # stap 3: inlezen featurescripts
    #
    #       3a: ophalen initialisatiebestand met macrovariabelen

    macrovars = {}
    with open(locatie_initscript) as inputfile:
        with open(locatie_parameterfile, 'w') as outputfile:
            for lines in inputfile:
                if "%LET" in lines.upper() and "=" in lines and lines.strip()[-1] == ";":
                    outputfile.write(lines)
                    macrovar = lines.split("=")[0].strip().split(" ")[-1]
                    waarde = lines.split("=")[1].split(";")[0].strip()

                    if "&" in waarde:  # macrovariabele vervangen door eerder gedefinieerde parameters
                        hulp1 = re.sub(r'\&(\w+)\.', r'{\1}', waarde)
                        hulp2 = re.sub(r'\&(\w+)\s', r'{\1} ', hulp1)
                        waarde = re.sub(r'\&(\w+.$)', r'{\1}', hulp2).format(**macrovars)
                    macrovars[macrovar] = waarde

    print('macrovariabelen aangemaakt')
    #        3b: verwerken van alle modulescripts

    for files in os.listdir(root_modulescripts):
        print(files)
        locatie_modulescript = root_modulescripts+"/"+files
        # to do regexp om versienummer op te halen en te controleren of naam voldoet
        if not (
                files.count("_") > 2 and
                "sim_k" in files and
                files.split("_")[-1][0] == "v" and
                files.endswith(".sas")):
            continue
        else:
            print("\n\nverwerk {}".format(files))

        moduleversie = files.split(".sas")[0]
        print('moduleversie:', moduleversie)
        modulenaam = "_".join(files.split("_")[2:]).split(".sas")[0]
        modulesettings = {}
        content = ""

        metadata_beschikbaar = True
        startblok = True
        with open(locatie_modulescript) as inputfile:
            for lines in inputfile:
                if startblok:
                    if ":" in lines:
                        tag = lines.split(":")[0].strip()
                        value = lines.split(":")[1].strip()
                        modulesettings[tag] = value
                    elif "REVISION HISTORY" in lines:
                        startblok = False
                else:
                    content += lines.strip()+" "
        if startblok:
            print("modulescript voldoet niet aan de eisen, geen revision history regel")
            continue
        if not metadata_beschikbaar:
            continue

        module = SimModule()
        module.verwerk_modulescript(modulesettings, moduleversie, macrovars=macrovars)
        tabelreferenties = {}
        for referentie in module.brontabellen:
            print(referentie)
            tabelref, tabelnaam = referentie.split(" ")
            # print(tabelnaam,tabelref)
            tabelreferenties[tabelref] = tabelnaam

        zonder_commentaar = re.sub(r"/\*[^*]*\*+(?:[^/*][^*]*\*+)*/", '', content.replace("\r", " ").replace("\n", " "))
        executequery = re.search(r"EXECUTE\s*\((.*?)\)\s*BY td", zonder_commentaar, re.IGNORECASE).group(1).strip()
        # codestuk tussen execute( en ) by td wordt opgehaald.
        # \s* staat een willekeurig aantal spaties toe tussen execute en (
        hulp1 = re.sub(r'\&(\w+)\.', r'{\1}', executequery)  # vervang &*. door {*}
        hulp2 = re.sub(r'\&(\w+)\s', r'{\1} ', hulp1)  # vervang onafgesloten &* door {*}
        zonder_commentaar = re.sub(r'\&(\w+.$)', r'{\1}', hulp2)  # vervang &*$ aan einde van regel door {*}

        # print(statement.format(**parameters))
        select_code = ""  # hierin komt de code waarin de kolommen gedefinieerd worden

########################################################################################################################
# haakjeslevel wordt bepaald om start en eindpunt te bepalen van codegedeelte waar kolommen in worden gedefinieerd
# dat is het stuk code tussen het eerste haakje open en de eerste from. De from mag geen deel uitmaken van een functie.
# de bedoelde from mag volgens de syntax nooit tussen haakjes staan en moet dus op zelfde niveau zitten.
# commentaar is al uit de code verwijderd, dus geen last van /* ...from...*/
#
# Level -1 wordt alleen gebruikt om het startpunt makkelijk te kunnen onderscheiden. Startpunt is overgang van -1 naar 1
# Eindpunt is from gedeelte op level 1
# create multiset table LIB.TAB as (select t1.k1, extract(year from k2) from LIB.TAB2 t1) with data primary index (finr)
# Level -1                         1                     2            1                 0                         1    0
#                                  *startpunt                         *eindpunt
# codedeel                         ____________________________________
#
# Tenslotte wordt de select eraf gehaald.
# select_code                              ____________________________
########################################################################################################################
        level = -1
        for index, characters in enumerate(zonder_commentaar.lower()):
            if level == -1:
                if characters == "(":
                    level += 2
            elif characters == "(":
                level += 1
            elif characters == ")":
                level -= 1
            elif level == 1:
                if characters == "f" and zonder_commentaar.lower()[index-1:index+5] == " from ":
                    codedeel = zonder_commentaar[0:index-1]
                    select_code = re.search(r"(?i)select(.*)", codedeel).group()[7:]

########################################################################################################################
# bepalen van de naam van de kolommen (stukje tussen de punt of laatste spatie en de komma)
# haakjeslevel wordt bepaald voor komma's in functies
# codedeel:        t1.k1, extract(year from k2) as k2, substr(k3,1,9) as k3
# haakjeslevel    0      1            0              1      0
#
# kolom               k1
# kolomcode        _____
# kolom                                            k2
# kolomcode              ____________________________
# kolom                                                                  k3
# kolomcode                                           _____________________
########################################################################################################################
        level = 0
        kolomcode = ""
        # print('selectcode:', select_code)

        for index, characters in enumerate(select_code):
            if level == 0 and characters == ",":
                kolom = kolomcode.strip().replace(" ", ".").split(".")[-1]
                kolomcode_copy = kolomcode
                for character in "*<>-+=/\\()|\t,":
                    kolomcode_copy = " "+kolomcode_copy.replace(character, " ")
                # verwijder alles tussen enkele quotes, hierbinnen staan geen bronkolommen.
                kolomcode_copy = re.sub(r"'.*?'", "''", kolomcode_copy).strip()

                # laatste stuk na spatie of punt geeft kolomnaam aan
                if not "{};{}".format(moduleversie, kolom) in featureset:
                    if kolom in ["finr", "peildatum"]:
                        pass
                    else:
                        print("feature {} uit script {} kan niet "
                              "worden gevonden in datadictionary".format(kolom, moduleversie))
                else:
                    featureset[moduleversie+";"+kolom].bronkolommen = \
                        list({word for word in kolomcode_copy.split(" ")
                              if ("." in word and word.split('.')[0].lower() in tabelreferenties and
                                  not word.replace('.', '').isdigit())})

                    module.kolommen.append(kolom)
                    featureset[moduleversie+";"+kolom].definitie = kolomcode
                    # print('kolomcode:', kolomcode)
                    featureset[moduleversie+";"+kolom].dependencies =\
                        list({word for word in kolomcode_copy.split(" ")[:-2] if word in module.kolommen})

                    print(modulenaam, kolom, 'dependencies', featureset[moduleversie+";"+kolom].dependencies,
                          'bronkolommen', featureset[moduleversie+";"+kolom].bronkolommen)
                    # print(tabelreferenties)
                    if featureset[moduleversie+";"+kolom].dependencies:
                        # print('bijbehorende kolomcode: ',kolomcode_copy.split(" ")[:-2])
                        pass
                    for word in kolomcode_copy.split(" "):
                        if word \
                                and word != "''" \
                                and word.lower() not in td_keywordlist \
                                and word not in featureset[moduleversie+";"+kolom].bronkolommen +\
                                        [kolom]+featureset[moduleversie+";"+kolom].dependencies \
                                and not word.replace('.', '').isdigit():
                            if "{" in word:
                                # print(featureset[moduleversie+";"+kolom].parameters)
                                if not word[1:-1]+" " in featureset[moduleversie+";"+kolom].parameters:
                                    print('A) kan token {word} niet plaatsen'.format(word=word))
                            else:
                                print(moduleversie, 'B) kan token {word} niet plaatsen'.format(word=word))
                                # print(kolomcode)
                kolomcode = ""
            else:
                kolomcode += characters
                if characters == "(":
                    level += 1
                elif characters == ")":
                    level -= 1

        # laatste kolom heeft geen komma aan het einde maar een spatie
        kolom = kolomcode.strip().replace(" ", ".").split(".")[-1]
        # print(kolomcode)
        kolomcode_copy = kolomcode
        for character in "*<>-+=*/\\()^|\t,":
            kolomcode_copy = " "+kolomcode_copy.replace(character, " ")
        # verwijder alles tussen enkele quotes, hierbinnen staan geen bronkolommen.
        kolomcode_copy=re.sub(r"'.*?'", "''", kolomcode_copy).strip()

        if not "{};{}".format(moduleversie, kolom) in featureset:
            if kolom in ["finr", "peildatum"]:
                pass
            else:
                print("feature {} uit script {} kan niet worden gevonden in datadictionary".format(kolom, moduleversie))
        else:
            module.kolommen.append(kolom)
            featureset[moduleversie+";"+kolom].definitie = kolomcode
            featureset[moduleversie+";"+kolom].bronkolommen=\
                list({word for word in kolomcode_copy.split(" ")
                      if ("." in word and word.split('.')[0].lower() in tabelreferenties and
                          not word.replace('.', '').isdigit())})
            featureset[moduleversie+";"+kolom].dependencies=\
                list({word for word in kolomcode_copy.split(" ")[:-2] if word in module.kolommen})

            print(modulenaam, kolom, 'dependencies', featureset[moduleversie+";"+kolom].dependencies, 'bronkolommen',
                  featureset[moduleversie+";"+kolom].bronkolommen)
            if featureset[moduleversie+";"+kolom].dependencies:
                        # print('bijbehorende kolomcode: ',kolomcode_copy.split(" ")[:])
                        pass
            for word in kolomcode_copy.split(" "):
                if word \
                        and word != "''" \
                        and word.lower() not in td_keywordlist \
                        and word not in featureset[moduleversie+";"+kolom].bronkolommen+[kolom] \
                        + featureset[moduleversie+";"+kolom].dependencies \
                        and not word.replace('.', '').isdigit():
                    if "{" in word:
                        # print(featureset[moduleversie+";"+kolom].parameters)
                        if not word[1:-1]+" " in featureset[moduleversie+";"+kolom].parameters:
                            print('C) kan token {word} niet plaatsen'.format(word=word))
                    else:
                        print(moduleversie, 'D) kan token {word} niet plaatsen'.format(word=word))
                        # print(kolomcode)

########################################################################################################################
# restant van de code ophalen en code opmaken.
# de naam dg_i_o_40ana_int.keytable_finr_validatie is hardcoded. Met deze tabel moet de tabel zijn aangemaakt.
# deze naam moet dus ook voorkomen in de tabeldefinitie
########################################################################################################################

        code_postfix = zonder_commentaar.split(select_code)[1]
        code_postfix_copy = code_postfix
        for character in "*<>-+=*/\\()^|\t,":
            code_postfix_copy = " "+code_postfix_copy.replace(character, " ")
        code_postfix_copy = re.sub(r"'.*?'", "''", code_postfix_copy).strip()
        module.bronkolommen = \
            list({word for word in code_postfix_copy.split(' ')
                  if ("." in word and word.split('.')[0].lower() in tabelreferenties
                      and not word.replace('.', '').isdigit())})
        print('module.bronkolommen', module.bronkolommen)

        for replacements in [["dg_i_o_40ana_int.keytable_finr_validatie ", "&inputlocatie. "],
                             ["dg_i_o_40ana_int.keytable_finr_beljaar_vld ", "&inputlocatie. "],
                             ["dg_i_o_40ana_int.keytable_1mln ", "&inputlocatie. "],
                             [" left join", "\n    LEFT JOIN"],
                             [" inner join", "\n    INNER JOIN"],
                             [" group by", "\n    GROUP BY"],
                             [" on ", "\n    ON "],
                             [" qualify", "\n    QUALIFY"],
                             [")with", ") WITH"],
                             [" with", "\nWITH"],
                             [' from ', ' FROM ']
                             ]:
            module.codepostfix = ireplace(code_postfix, replacements[0], replacements[1])

        print(module.modulegroep, module.naam, module.key, "bevat volgende kolommen:")
        print(module.kolommen)

        if moduleversie not in modulelijst:
            print(moduleversie, "kan niet worden gekoppeld aan moduleversie in datadictionary")
        else:
            locatie_keygroep = '{}/{}'.format(locatie_keymodules, module.key)
            locatie_modulegroep = locatie_keygroep+'/'+module.modulegroep
            locatie_moduleversie = locatie_modulegroep+'/'+module.naam
            if not os.path.exists(locatie_keygroep):
                # print(locatie_keygroep, 'bestaat nog niet, wordt nu aangemaakt')
                os.makedirs(locatie_keygroep)
            else:
                pass
                # print(locatie_keygroep, 'bestaat al')

            if not os.path.exists(locatie_modulegroep):
                # print(locatie_modulegroep, 'bestaat nog niet, wordt nu aangemaakt')
                os.makedirs(locatie_modulegroep)
            else:
                pass
                # print(locatie_modulegroep, 'bestaat al')

            if not os.path.exists(locatie_moduleversie):
                # print(locatie_moduleversie, 'bestaat nog niet, wordt nu aangemaakt')
                os.makedirs(locatie_moduleversie)

                with open(locatie_moduleversie+'.json', 'w') as outputfile:
                    # hier nog allerlei modulespecifieke attributen toevoegen
                    outputfile.write(module.schrijfweg())
            else:
                pass
                # print(locatie_moduleversie, 'bestaat al')

            for kolom in module.kolommen:
                # p rint('verwerk kolom',kolom)
                if moduleversie+";"+kolom in featureset:
                    with open(locatie_moduleversie+"/"+kolom+'.json', 'w') as outputfile:
                        # print('output weggeschreven naar:', locatie_moduleversie+"/"+kolom+'.py')
                        outputfile.write(featureset[moduleversie+";"+kolom].schrijfweg())
                else:
                    print(moduleversie+";"+kolom, 'niet in featureset')
    return




