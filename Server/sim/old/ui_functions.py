"""Deze module bevat de functies die onder de Interface-buttons hangen.

.. todo:: * Controleren of keymodule na init interface nog ergens aangeroepen wordt \n\n \
* In de interface worden custom features toegevoegd, maar deze zitten niet in de beschikbare features van de keymodules.


"""

from settings import *
import os
from openpyxl import Workbook
#from .tdquery import *
from ui_settings import *
from lineage import lineage_total
import json
from sharedclasses import Feature


def on_load_selectie_button_clicked(self):
    """Zorgt ervoor dat de feature-bestanden (als json) vanuit de opgegeven directory ingeladen worden, waarbij
    de features in de interface aangevinkt worden.
    """
    with self.configuratie.interface.elements['saveload_logging']:
        clear_output()

        selectiondirectory = self.configuratie.interface.elements['lineedit_selectieroot_text'].value +\
                             self.configuratie.interface.elements['selectiefolder_dropdown'].value
        #print("load clicked",selectiondirectory)
        for jsonfile in os.listdir(selectiondirectory):
            #print(jsonfile)
            if jsonfile.endswith(".json"):
                feature_id = jsonfile.split(".json")[0]
                #print(feature_id)
                moduleversie, featurenaam = feature_id.split(";")
                with open(selectiondirectory+"/{}.json".format(feature_id)) as infile:
                    json_input = json.load(infile)
                if feature_id not in self.configuratie.interface.featuredict:
                    feature = Feature(featuredict=json_input) #let op, hierbij wordt de keymodule niet aangepast!
                    self.configuratie.interface.moduleversiedict[moduleversie].beschikbare_features.append(feature_id)
                    self.configuratie.interface.add_feature(self.configuratie.interface.moduleversiedict[moduleversie],feature)
                    print(feature_id,"toegevoegd aan {}".format(moduleversie))
                else:
                    print(feature_id,"let op! Deze feature was al beschikbaar en is niet opnieuw aangemaakt.")
        self.configuratie.interface.build()
        for jsonfile in os.listdir(selectiondirectory):
            if jsonfile.endswith(".json"):
                feature_id = jsonfile.split(".json")[0]
                self.configuratie.interface.elements['{}_checkbox'.format(feature_id)].value = True
                print(feature_id,"checked")


def on_delete_selectie_button_clicked(self):
    """Geeft als output het linux-commando om de opgegeven directory te verwijderen.
    """
    with self.configuratie.interface.elements['saveload_logging']:
        clear_output()
        filelocatie = self.configuratie.interface.elements['lineedit_selectieroot_text'].value + \
                      self.configuratie.interface.elements['selectiefolder_dropdown'].value

        print("verwijder handmatig met $ rm -rf {}".format(filelocatie))
        # os.remove(filelocatie)

        #self.configuratie.interface.elements['selectiefolder_dropdown'].options = os.listdir(
        #    self.configuratie.interface.elements['lineedit_selectiefolder_text'].value)
   

def on_save_selectie_button_clicked(self):
    """Zorgt ervoor dat de aangevinkte features in de interface weggeschreven worden (als json-bestanden) in de
    opgegeven directory.
    """
    with self.configuratie.interface.elements['saveload_logging']:
        clear_output()
        print('wegschrijven van features')
        actieve_features = []
        outputfolderlocatie = self.configuratie.interface.elements['lineedit_selectieroot_text'].value + \
                              self.configuratie.interface.elements['lineedit_selectiefolder_text'].value
        if not os.path.exists(outputfolderlocatie):
            os.makedirs(outputfolderlocatie)

            for items in self.configuratie.interface.available_features:
                if self.configuratie.interface.elements[items[1].id + "_checkbox"].value:
                    print(items[1].id)
                    with open(outputfolderlocatie + '/{}.json'.format(items[1].id),'w') as outputfile:
                        outputfile.write(items[1].schrijfweg()+"\n")

            self.configuratie.interface.elements['selectiefolder_dropdown'].options = os.listdir(
                    self.configuratie.interface.elements['lineedit_selectieroot_text'].value)
        else:
            print("deze folder bestaat al, geef een andere naam op.")

        #with open(self.configuratie.interface.elements['lineedit_selectieroot_text'].value +
        #                  self.configuratie.interface.elements['lineedit_selectiefile_text'].value, 'w') as outputfile:
        #    for items in self.configuratie.interface.available_features:
        #        if self.configuratie.interface.elements[items[1].id + "_checkbox"].value:
        #            actieve_features.append(items[1].id)
        #    outputfile.write("\n".join(actieve_features))
        #    print(actieve_features)
        #    self.configuratie.interface.elements['selectiefolder_dropdown'].options = os.listdir(
        #        self.configuratie.interface.elements['lineedit_selectiefolder_text'].value)


def on_export_datadictionary_button_clicked(self):
    """ ..todo:: @Wouter: verwijderen? Wordt nu niet gebruikt.

    """
    moduleversie_exportkolommen = ['eigenaar',
                                   'reviewer',
                                   'auteur',
                                   'modulegroep']
    feature_exportkolommen = ['naam',
                              'bronkolommen',
                              'moduleversie',
                              'reviewed_by_peer',
                              'numerical_character',
                              'td_format',
                              # 'type_variable',
                              'length',
                              'format',
                              'label',
                              'sample_values',
                              'beschrijving',
                              'opmerkingen',
                              'betekenis_missing',
                              'betekenis_zero',
                              'betekenis_negatief',
                              'verversfrequentie',
                              'kwaliteit_databron',
                              'doelbinding',
                              # 'featuredepency',
                              # 'moduledependency',
                              'parameters',
                              'definitie']

    with self.configuratie.interface.elements['out_logging']:
        clear_output()
        print(self.configuratie.interface.elements['lineedit_export_datadictionary'].value)

        dest_filename = self.configuratie.interface.elements['lineedit_export_datadictionary'].value
        wb = Workbook()
        ws1 = wb.active
        ws1.title = 'Datadictionary'

        rownr = 1
        for colnr, attrib in enumerate(feature_exportkolommen):
            ws1.cell(column=colnr + 1, row=rownr, value=attrib)
        for index, attrib in enumerate(moduleversie_exportkolommen):
            ws1.cell(column=colnr + 1 + index + 1, row=rownr, value=attrib)

        for items in self.configuratie.interface.available_features:
            if self.configuratie.interface.elements[items[1].naam + "_checkbox"].value:
                rownr += 1
                for colnr, attrib in enumerate(feature_exportkolommen):
                    ws1.cell(column=colnr + 1, row=rownr, value=vars(items[1])[attrib])
                for index, attrib in enumerate(moduleversie_exportkolommen):
                    ws1.cell(column=colnr + 1 + index + 1, row=rownr, value=vars(items[0])[attrib])
        wb.save(filename=dest_filename)


def on_generate_keytable_npbelastingplichtig_button_clicked(self):
    """Genereer een keytable van NP belastingplichtigen in de opgegeven locatie op de opgegeven peildatum.
    """
    with self.configuratie.interface.elements['out_logging']:
        clear_output()
        print("genereer keytable {} ...".format(self.configuratie.keytable))
        execute_query("drop table {}".format(self.configuratie.keytable), ignore=True)
        execute_query(definitie_npbelastingplichtig.format(keytabel=self.configuratie.keytable,
                                                           peildatumyyyymmdd=self.configuratie.peildatum),
                      ignore=True)
        print('keytable klaar voor gebruik')
    pass


def on_generate_keytable_nnpactief_button_clicked(self):
    """Genereer een keytable van actieve NNP belastingplichtigen in de opgegeven locatie op de opgegeven peildatum.
    """
    with self.configuratie.interface.elements['out_logging']:
        print("genereer keytable {} ...".format(self.configuratie.keytable))
        execute_query("drop table {}".format(self.configuratie.keytable), ignore=True)
        execute_query(definitie_nnpactief.format(keytabel=self.configuratie.keytable,
                                                 peildatumyyyymmdd=self.configuratie.peildatum),
                      ignore=True)
        print('keytable klaar voor gebruik')

    pass

########################################################################################################################
# onderstaand stukje code nog even opschonen. +="""...""" beter vervangen door ''.join-statements.
# bovendien hoeft maar een keer geloopt te worden door de actieve_features.
# ook aanpassen op key! nu standaard finr, peildatum
# nu wordt gebruik gemaakt van de macrovariabele keytable_finr_validatie binnen alle modules.
# mogelijk is de naam keytable_k1_validatie beter. Deze moet dan worden toegevoegd.
########################################################################################################################


def on_create_abt_code_button_clicked(self):
    """Genereer de ABT code o.b.v. de aangevinkte features en opgegeven locatie, en schrijf deze naar de output.
    """
    # PARAMS = ACTIEVE FEATURES 
    # VOEG ACTIEVE FEATURES TOE
    with self.configuratie.interface.elements['codegenerator_out_logging']:
        clear_output()
        actieve_features = {}
        statements = []
        lineage = {}  # {datagebied:{tabel:kolom}}
        for items in self.configuratie.interface.available_features:
            if self.configuratie.interface.elements[items[1].id + "_checkbox"].value:
                actieve_features.setdefault(items[0], []).append(items[1])

        # het joinstatement wordt gemaakt, voordat de dependencies worden bepaald.
        # zo worden dependent variabelen niet meegenomen in de uiteindelijke output.
        # in het stuk code onder het statement worden kolommen toegevoegd aan de actieve features
        # aan de hand van de dependencies.

        joinstatement_start = """
CREATE MULTISET TABLE {keytable} AS (""".format(keytable=self.configuratie.abt)
        joinstatement_select = """
SELECT
     t0.finr
     ,t0.peildatum"""
        for tx, module in enumerate(actieve_features):
            for feature in actieve_features[module]:
                joinstatement_select += "\n     ,t{index}.{kolomnaam}".format(index=tx + 1, kolomnaam=feature.naam)

        joinstatement_from = """
\nFROM {} t0""".format(self.configuratie.keytable)
        for tx, module in enumerate(actieve_features):
            joinstatement_from += """\nLEFT JOIN v_{tabelnaam} t{index}
            ON t0.finr = t{index}.finr AND t0.peildatum = t{index}.peildatum""".format(
                tabelnaam=module.naam, index=tx + 1)
        joinstatement_end = """
) WITH DATA PRIMARY INDEX(FINR, PEILDATUM)
"""
        joinstatement = ''.join([joinstatement_start, joinstatement_select, joinstatement_from, joinstatement_end])

        inputmodules = list(actieve_features.keys())
        for module in inputmodules:
            # print(module.naam)
            tabelrefs = {}
            for ref, tabellocatie in module.brontabellen.items():
                # print(tabellocatie)
                tabelrefs[ref] = tabellocatie
                datagebied, tabelnaam = tabellocatie.split(".")
                lineage.setdefault(datagebied, {})
                lineage[datagebied].setdefault(tabelnaam, set())
                # dependencies regelen:
                if tabelnaam.startswith('sim_') and not self.configuratie.interface.moduleversiedict[
                    tabelnaam] in inputmodules:
                    # moet nog worden aangepast als versie is aangepast
                    inputmodules.append(self.configuratie.interface.moduleversiedict[tabelnaam])
                    # print('module toegevoegd',tabelnaam)
            # statements moeten nog in juiste volgorde worden afgehandeld
            statement = "create volatile table v_{tabelnaam}".format(
                tabelnaam=module.naam) + " AS(\n/*"+ module.beschrijving + "*/\n\tselect \n\tt0.finr,\n\tt0.peildatum"

            for kolomlocatie in module.bronkolommen:  # moduleafhankelijke kolommen
                tabelref, kolomnaam = kolomlocatie.split('.')
                datagebied, tabelnaam = tabelrefs[tabelref].split(".")
                lineage[datagebied][tabelnaam].add(kolomnaam)

            # onderstaande kan efficienter...
            if module in actieve_features:
                inputkolommen = actieve_features[module]
            else:
                inputkolommen = []

            for feature in inputkolommen:  # feature-afhankelijke kolommen

                for kolomlocatie in feature.bronkolommen:
                    tabelref, kolomnaam = kolomlocatie.lower().split('.')
                    # print(datagebied,tabelnaam,tabelref)
                    datagebied, tabelnaam = tabelrefs[tabelref].split(".")
                    lineage[datagebied][tabelnaam].add(kolomnaam)
                for columndependency in feature.dependencies:
                    # print(feature.id, columndependency)
                    # als dit het geval is, moet ook de oorspronkelijke feature worden toegevoegd.
                    # Dit gebeurt op deze manier iteratief. Wanneer een feature afhankelijk is van een andere
                    # feature die weer afhankelijk is van een andere feature gaat het ook goed.
                    dependent_feature = self.configuratie.interface.featuredict[
                        feature.id.replace(feature.naam, columndependency)]
                    if not dependent_feature in actieve_features[module]:
                        inputkolommen.append(dependent_feature)

                # parameteroplossing is aangepast bij het genereren van de custom features.
                parameters = {}
                #if feature.id in self.configuratie.interface.parameterwidgets:
                #    parameterwidgets = self.configuratie.interface.parameterwidgets[feature.id]
                #    for parameter, parameterwidgetnaam in parameterwidgets:
                #        paramval = self.configuratie.interface.elements[parameterwidgetnaam].value
                #        if type(paramval) is list or type(paramval) is tuple:
                #            parameters[parameter] = ",".join(paramval)
                #        else:
                #            parameters[parameter] = paramval

                statement += "\n\n\t," + "/*" + feature.beschrijving + "*/ \n" + feature.definitie.format(**parameters)


            statement += "\n"+module.codepostfix.replace("FROM","\nFROM").replace("LEFT JOIN","\nLEFT JOIN") + " ON COMMIT PRESERVE ROWS\n\n"
            statements.append(statement.replace("{keytable_finr_validatie}", self.configuratie.keytable))
            # print(lineage)
            # print(statement)

        # print(joinstatement)
        statements.append(joinstatement)

        # print('benodige kolommmen en tabellen:')

        self.configuratie.abt_statements = statements
        for datagebied, tabelnaam in lineage.items():
            # print(datagebied)
            for tabelnaam, kolommen in lineage[datagebied].items():
                for kolomnaam in kolommen:
                    # print('\t{tabelnaam}.{kolomnaam}'.format(tabelnaam=tabelnaam, kolomnaam=kolomnaam))
                    pass

        print("\n".join(statements))
    return


def on_save_abt_code_button_clicked(self):
    """Sla de ABT code o.b.v. de aangevinkte features op in de opgegeven directory.
    """
    with self.configuratie.interface.elements['codegenerator_out_logging']:
        clear_output()
        abt_statements = self.configuratie.abt_statements[:]
        # copy, aanpassing wordt gemaakt door macrovariabelen te vertalen voor sas

        current_timestamp = str(datetime.now()).replace(" ", "_").replace("-", "").replace(":", "").split(".")[0]
        outputlocatie = "{}{}_{}.sas".format(self.configuratie.folder_sasscript_abtcode,
                                             self.configuratie.abt.split(".")[-1],
                                             current_timestamp)
        # bepalen libnames

        librefs = []
        dubrefs = []
        refs = [ref.lower() for ref in schemas.values()]
        for key, val in schemas.items():
            if refs.count(val.lower()) == 1 and key != 'singlequote':
                for statement in abt_statements:

                    if val in statement and [key, val] not in librefs:
                        librefs.append([key, val])
                    statement = ireplace(statement, " " + val, " &" + key + ".")

            elif refs.count(val.lower()) > 1 and key != 'singlequote':
                for statement in abt_statements:

                    if val in statement:
                        dubrefs.append(val)

        for items in librefs:
            pass
            # print(items)
            # print('{} vervangen door {}'.format(items))

        for items in set(dubrefs):
            #print('{} kan niet eenduidig vertaald worden'.format(items))
            #verplaatsen naar testwerk
            pass

        with open(outputlocatie, 'w') as outputfile:
            outputfile.write("/* Code weggeschreven op {} door {}*/\n".format(current_timestamp, username))
            outputfile.write("""

%LET teradata_connect_string = 
    SERVER="tdata03" 
    AUTHDOMAIN=TeradataAuth
    TPT=yes
    DEFER=yes
    CONNECTION=global
    BULKLOAD=yes
    mode=Teradata;
    
/* Set libraries & TERADATA options */
option msglevel              = n nostsuffix
SAStrace                     = ',,,ds' SAStraceloc=SASlog
sql_ip_trace                 = source sqlgeneration=dbmust
sqlmapputto                  = SAS_put dbidirectexec 
MPRINT MPRINTNEST MLOGIC MLOGICNEST SYMBOLGEN MERROR;

libname x teradata authdomain=teradataauth server=tdata03 mode=teradata connection=global;

""")

            for items in librefs:
                outputfile.write('%LET {} = {};\n'.format(items[0], items[1]))
            for items in set(dubrefs):
                outputfile.write('/*{} kon niet eenduidig vertaald worden naar een macrovariabele*/\n'.format(items))

            for statement in abt_statements:
                outputfile.write(sassify(statement) + "\n\n")

        print("code weggeschreven naar {}".format(outputlocatie))
        return


def on_run_abt_code_button_clicked(self):
    """Run de ABT code o.b.v. de opgegeven keytable en opgegeven directory-locatie van de ABT.
    """
    with self.configuratie.interface.elements['codegenerator_out_logging']:
        clear_output()
        if self.configuratie.abt_statements:
            print('uitvoeren abt_code...')
            execute_query("drop table {}".format(self.configuratie.abt), ignore=True)
            execute_queries(self.configuratie.abt_statements)
            print('code uitgevoerd!')
        else:
            print('geen code gegenereerd')
        return


def on_check_autorisatie_button_clicked(self):
    """Check of de gebruiker de benodigde autorisaties heeft (o.b.v. opgegeven gebruik (bv. Pilot) en opgegeven rol
    (bv. Ontwikkelaar Algemeen)) om de aangevinkte features te genereren. In de logging verschijnt of de gebruiker
    deze autorisaties heeft, of dat er tabellen zijn waar de gebruiker geen toegang (meer) toe heeft.
    """
    with self.configuratie.interface.elements['out_logging']:
        clear_output()

        #execute_query("sel * from DG_I_O_40ANA_SIM.PROJECTMAPPING")
        #query_result = print_query("sel * from DG_I_O_40ANA_SIM.PROJECTMAPPING")


        #1. gewenste informatie (o.b.v. aangeclickte features in interface

        lineage = lineage_total(self.configuratie)

        #print("lineage:\n")
        #print(lineage)
        #print("\nlineage einde\n")
        #opbouw van lineage:
        # {'DG_I_P_30INF_LOON': {'i_loon_werknemer': {'loon_bruto_inkomen_bdr', 'flg_alo_lb_jr'}},
        # 'DG_I_O_40ANA_INT': {'keytable_1mln': {'finr', 'peildatum'}}}
        # {DATAGEBIED: {TABEL: {KOLOM}}}

        #for datagebied, tabelnaam in lineage.items():
        #    print(datagebied)
        #    for tabelnaam, kolommen in lineage[datagebied].items():
        #        for kolomnaam in kolommen:
        #            print('\t{tabelnaam}.{kolomnaam}'.format(tabelnaam=tabelnaam, kolomnaam=kolomnaam))

        #2. informatie waartoe je gerechtigd bent (o.b.v. gebruik en rol) op huidige tijdstip
        # dit is een list []
        #justified_data = query2table("SELECT * FROM DG_I_O_40ANA_SIM.TABELMAPPING")

        #de waarde van het type gebruik (bv. 'Pilot')
        selected_gebruik = self.configuratie.interface.elements['doelbinding_gebruik_select'].value
        selected_rol = self.configuratie.interface.elements['doelbinding_rol_select'].value

        #in toekomst kunnen we beter de naam in de interface mappen op de naam zoals gebruikt in tabel DATAGEBRUIK
        # d.w.z. Pilot -> PIL (ipv oplossen in de code)
        def gebruik_switcher(x):
            """Mapt de naamgeving van het Gebruik in de interface op de naamgeving in de Datagebruik-tabel.

            :param x: naam van het Gebruik in de Interface
            :return: naam van het Gebruik in de Datagebruik-tabel
            """
            return {
                'Pilot': 'PIL',
                'Productie': 'PROD',
            }.get(x, None)

        def rol_switcher(x):
            """Mapt de naamgeving van de Rol in de interface op de naamgeving in de Datagebruik-tabel.

            :param x: naam van de Rol in de Interface
            :return: naam van de Rol in de Datagebruik-tabel
            """
            return {
                'Ontwikkelaar IH Niet Winst': 'Ontwikkelaar IH Niet Winst',
                'Ontwikkelaar IH Winst': 'Ontwikkelaar IH Winst',
            }.get(x, None)

        #Gebruik (Pilot), Rol (Ontwikkelaar IH Niet Winst) worden in de query verwerkt, zodat het in de database wordt opgelost (en
        # de db niet verlaat)
        #nu gekozen voor: in LAB mag alles (justified_data levert een lege lijst op, maar daar hoeft
        # niet naar gekeken te worden); in Pilot of productie niet. Pilot en productie leveren verschillende regels op.
        #Voor IH Niet Winst zit in de DATAGEBRUIK-tabel ook een regel voor productie die dus niet geselecteerd wordt voor pilot
        #uit query gehaald omdat deze enkel gebruikt worden om te filteren: t1.GEBRUIK_TYPE, t2.PROJECTNAAM, , t1.GOEDKEURING_VANAF_TS, t1.GOEDKEURING_TOT_TS
        #Je levert op: datagebied, tabelnaam, current_status (deze geeft aan of het record nu nog geldig is (1) of niet (0). Door ook
        # records op te leveren die niet meer geldig zijn, kunnen we dat terugkoppelen naar de gebruiker
        # ("De autorisaties voor ... zijn niet meer actief")
        # GROUP: uniek op [datagebied, tabelnaam] waarbij current_status op 1 staat indien er een actieve status is (er kunnen ook inactieve
        #  statussen zijn, maar die worden dan genegeerd)
        print("Uitvoeren van query op Teradata ...")
        justified_data = query2table("""
            SELECT
                t3.DATAGEBIED
                , t3.TABELNAAM
                , MAX(CASE WHEN t1.GOEDKEURING_VANAF_TS <= CURRENT_TIMESTAMP AND t1.GOEDKEURING_TOT_TS > CURRENT_TIMESTAMP THEN 1 ELSE 0 END) AS CURRENT_STATUS
            FROM DG_I_O_40ANA_SIM.DATAGEBRUIK t1
            INNER JOIN (SELECT * FROM DG_I_O_40ANA_SIM.PROJECTMAPPING) t2
            ON t1.VAN_DATA_ID = t2.DATA_ID
            INNER JOIN (SELECT * FROM DG_I_O_40ANA_SIM.TABELMAPPING) t3
            ON t1.NAAR_DATA_ID = t3.DATA_ID
            /* selecteer geldige records */
            WHERE
                t1.VAN_DATA_GELDIG_VANAF_TS <= CURRENT_TIMESTAMP AND t1.NAAR_DATA_GELDIG_VANAF_TS <= CURRENT_TIMESTAMP
                AND t1.GEBRUIK_TYPE = '{gebruik}'
                AND t2.PROJECTNAAM = '{rol}'
            GROUP BY t3.DATAGEBIED, t3.TABELNAAM

        """.format(gebruik=gebruik_switcher(selected_gebruik), rol=rol_switcher(selected_rol)), log=False)
        print("Query uitgevoerd")
        #print("justified data:\n")
        #print(justified_data)

        #TODO: wellicht nog een check bij Marc vd Valk van geldigheid- en goedkeuring-vars (van het gehele gebruik van de tabellen)


        #huidige ouput IHNW:
        #module sim_k1_bank_beleggen_v1 , tabel bank_beleg_bdr_custom1
        #module sim_k1_clc_burgerlijkestaat_v1 , tabel clc_burgerlijke_staat_cat

            #lineage: {'DG_I_P_30INF_BANK': {'i_bank_np_beleggen': {'rbg_rek_gezamenlbel', 'rbg_totaal_saldo_eur', 'rbg_begin_d', 'rbg_eind_d', 'finr'}},
            # 'DG_I_P_30INF_CLC': {'g_clc_np': {'bvr_pnp_verval_d', 'bvr_pso_ovl_d', 'finr'}, 'g_clc_finr_relaties': {'bvr_pre_afvoer_d', 'bvr_pre_finr2',
            # 'bvr_pre_verval_d', 'bvr_pre_relsoort', 'bvr_pre_opvoer_d', 'bvr_pre_ingang_d', 'finr'}}, 'DG_I_O_40ANA_INT': {'keytable_1mln': {'peildatum', 'finr'}}}

            #justified_data (IHNW, Pilot):
            #[['DG_I_O_30INF_CLC', 'g_clc_np', Decimal('1')], ['DG_I_O_40ANA_INT', 'keytabel_1mln', Decimal('1')],
            # ['DG_I_P_30INF_BANK', 'i_bank_np_beleggen', Decimal('1')],
            # ['DG_I_O_30INF_CLC', 'g_clc_finr_relaties', Decimal('1')]]

        #{DG_I_O_30INF_CLC: {g_clc_finr_relaties: 1, g_clc_finr_relaties: 0}}

        #huidige ouput IHW:
        #module sim_k1_clc_relaties_v1 , tabel clc_entiteitnr


        #zoek alle gebruikte datagebied.tabellen uit lineage op in justified_data
        #not_accessible: tabellen die niet accessible zijn op dit moment (incl. pas verlopen autorisaties)
        not_accessible = []
        #previously_accessible: tabellen die ooit accessible zijn geweest, maar nu niet meer (zijn ook opgenomen in not_accessible)
        previously_accessible = []
        #iterate over datagebied keys
        for datagebied in lineage:
            for tabelnaam in lineage[datagebied]:
                found = 0

                for dg_table_status in justified_data:
                    #print("{} {}".format(datagebied, tabelnaam))
                    #print("testing {} {}".format(dg_table_status[0], dg_table_status[1]))
                    if datagebied == dg_table_status[0] and tabelnaam == dg_table_status[1]:
                        #print("found {} {}".format(datagebied, tabelnaam))

                        found = 1
                        if dg_table_status[2] == 0:
                            #print('0-loop')
                            previously_accessible.append([datagebied, tabelnaam])
                            not_accessible.append([datagebied, tabelnaam])
                        #omdat [datagebied, tabelnaam] uniek is en je hebt hem gevonden -> break de for-loop dg_table_status (niet verder zoeken)
                        break

                #TODO: gebruik een escape voor LAB, zodat alles accessible is (voor LAB is justified_data nu leeg, waardoor niks
                # accessible is)
                if found == 0:
                    not_accessible.append([datagebied, tabelnaam])

        print("\nJe hebt geen autorisaties tot de volgende tabellen\n(die worden gebruikt in de geselecteerde features):")
        not_accessible_str = ''
        previously_accessible_str = ''
        for i in not_accessible:
            not_accessible_str += '- {dg}.{tabelnaam}\n'.format(dg=i[0], tabelnaam=i[1])
        print(not_accessible_str)
        print("\nVan de volgende tabellen zijn de autorisaties verlopen:")
        for i in previously_accessible:
            previously_accessible_str += '- {dg}.{tabelnaam}\n'.format(dg=i[0], tabelnaam=i[1])
        print(previously_accessible_str)


        #result = print_query("SELECT * FROM DG_I_O_40ANA_SIM.PROJECTMAPPING")
        #execute_query("drop table DG_I_O_40ANA_SIM.mfv")
        return
